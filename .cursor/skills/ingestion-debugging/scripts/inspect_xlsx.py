#!/usr/bin/env python3
"""Inspect the raw structure of an Excel file.

Usage:
    uv run python .cursor/skills/ingestion-debugging/scripts/inspect_xlsx.py <file.xlsx>
"""

from __future__ import annotations

import json
import sys
from pathlib import Path


def inspect(path: Path) -> dict:
    try:
        import openpyxl
    except ImportError:
        print("openpyxl not available — trying xlrd for .xls files")
        import xlrd  # type: ignore[import]
        return _inspect_xls(path, xlrd)

    wb = openpyxl.load_workbook(path, data_only=True)
    result = {"file": path.name, "sheets": []}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Find header row — first row with 3+ non-None cells
        header_idx = 0
        for i, row in enumerate(rows):
            non_null = sum(1 for c in row if c is not None)
            if non_null >= 3:
                header_idx = i
                break

        header_row = rows[header_idx] if rows else []
        headers = [str(c) if c is not None else "" for c in header_row]
        data_rows = rows[header_idx + 1 : header_idx + 4]

        sheet_info = {
            "name": sheet_name,
            "total_rows": ws.max_row,
            "total_cols": ws.max_column,
            "header_row_index": header_idx,
            "headers": headers,
            "sample_rows": [
                [str(c) if c is not None else "" for c in row]
                for row in data_rows
            ],
        }

        # Merged cells
        merged = [str(m) for m in ws.merged_cells.ranges]
        if merged:
            sheet_info["merged_cells"] = merged

        result["sheets"].append(sheet_info)

    return result


def _inspect_xls(path: Path, xlrd) -> dict:  # type: ignore[no-untyped-def]
    wb = xlrd.open_workbook(str(path))
    result = {"file": path.name, "sheets": []}

    for sheet_name in wb.sheet_names():
        ws = wb.sheet_by_name(sheet_name)
        rows = [ws.row_values(i) for i in range(ws.nrows)]

        header_idx = 0
        for i, row in enumerate(rows):
            if sum(1 for c in row if c) >= 3:
                header_idx = i
                break

        headers = [str(c) for c in rows[header_idx]] if rows else []
        data_rows = rows[header_idx + 1 : header_idx + 4]

        result["sheets"].append({
            "name": sheet_name,
            "total_rows": ws.nrows,
            "total_cols": ws.ncols,
            "header_row_index": header_idx,
            "headers": headers,
            "sample_rows": [[str(c) for c in r] for r in data_rows],
        })

    return result


def main() -> None:
    if len(sys.argv) < 2:
        print("Usage: inspect_xlsx.py <file.xlsx>")
        sys.exit(1)

    path = Path(sys.argv[1])
    if not path.exists():
        print(f"File not found: {path}")
        sys.exit(1)

    result = inspect(path)

    for sheet in result["sheets"]:
        print(f"\n{'='*60}")
        print(f"Sheet: {sheet['name']}  ({sheet['total_rows']} rows × {sheet['total_cols']} cols)")
        print(f"Header row index: {sheet['header_row_index']}")
        print(f"\nHeaders ({len(sheet['headers'])}):")
        for i, h in enumerate(sheet["headers"]):
            print(f"  [{i}] {h!r}")
        print(f"\nSample data ({len(sheet['sample_rows'])} rows):")
        for row in sheet["sample_rows"]:
            print(f"  {row}")
        if sheet.get("merged_cells"):
            print(f"\nMerged cells: {sheet['merged_cells']}")

    print(f"\n{'='*60}")
    print("JSON:")
    print(json.dumps(result, indent=2, default=str))


if __name__ == "__main__":
    main()
