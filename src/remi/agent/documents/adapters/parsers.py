"""File parsers for the ingestion pipeline.

Converts raw file bytes into a ``Document`` (tabular rows + column names).
This is the inbound I/O boundary for uploaded files — all file-format
knowledge lives here and nowhere else.

Public surface:
  parse_document(filename, content, content_type) -> Document
      Single entry point.  Detects format from content_type + extension and
      delegates to the appropriate parser.  Raises ValueError for unsupported
      types — callers should not branch on content_type themselves.

  parse_csv / parse_excel
      Exposed for direct use in tests and batch loading.
"""

from __future__ import annotations

import csv
import datetime as _dt
import io
import re
import uuid
from typing import Any

from remi.agent.documents.adapters.text_parsers import (
    parse_docx,
    parse_image,
    parse_pdf,
    parse_text,
)
from remi.agent.documents.types import DocumentContent

_GENERIC_METADATA_PATTERNS = (
    r"exported\s+on",
    r"date\s+range",
    r"filter",
    r"exclude",
    r"include",
    r"base\s+report",
    r"as\s+of",
    r"level\s+of\s+detail",
    r"report\s+builder",
)

_MIN_HEADER_CELLS = 3

_CSV_CONTENT_TYPES = frozenset({"text/csv", "application/csv"})
_EXCEL_CONTENT_TYPES = frozenset(
    {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    }
)
_PDF_CONTENT_TYPES = frozenset({"application/pdf"})
_DOCX_CONTENT_TYPES = frozenset(
    {
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    }
)
_TEXT_CONTENT_TYPES = frozenset({"text/plain", "text/markdown"})
_IMAGE_CONTENT_TYPES = frozenset({"image/jpeg", "image/png", "image/gif", "image/webp"})


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------


def parse_document(
    filename: str,
    content: bytes,
    content_type: str,
    *,
    extra_skip_patterns: tuple[str, ...] = (),
    section_labels: frozenset[str] = frozenset(),
) -> DocumentContent:
    """Parse uploaded file bytes into a Document.

    Format is determined by content_type first, filename extension second.
    Raises ``ValueError`` for unsupported types so callers never need to
    branch on content_type themselves.

    *extra_skip_patterns* are domain-specific regex patterns for report
    header rows that should be skipped (passed through from DomainProfile).
    """
    ct = content_type.lower()
    name = filename.lower()

    # Tabular formats
    if ct in _CSV_CONTENT_TYPES or name.endswith(".csv"):
        return parse_csv(filename, content)
    if ct in _EXCEL_CONTENT_TYPES or name.endswith((".xlsx", ".xls")):
        return parse_excel(
            filename,
            content,
            extra_skip_patterns=extra_skip_patterns,
            section_labels=section_labels,
        )

    # Text formats
    if ct in _PDF_CONTENT_TYPES or name.endswith(".pdf"):
        return parse_pdf(filename, content)
    if ct in _DOCX_CONTENT_TYPES or name.endswith(".docx"):
        return parse_docx(filename, content)
    if ct in _TEXT_CONTENT_TYPES or name.endswith((".txt", ".md")):
        return parse_text(filename, content, ct or "text/plain")

    # Image formats
    if ct in _IMAGE_CONTENT_TYPES or name.endswith((".jpg", ".jpeg", ".png", ".gif", ".webp")):
        return parse_image(filename, content, ct or "image/jpeg")

    raise ValueError(
        f"Unsupported file type: {content_type!r}. "
        "Supported formats: CSV, Excel, PDF, Word, text, and images."
    )


# ---------------------------------------------------------------------------
# Format-specific parsers
# ---------------------------------------------------------------------------


def parse_csv(filename: str, content: bytes | str) -> DocumentContent:
    """Parse a CSV file into a DocumentContent with typed rows."""
    if isinstance(content, bytes):
        content = content.decode("utf-8-sig")

    reader = csv.DictReader(io.StringIO(content))
    columns = reader.fieldnames or []
    rows: list[dict[str, Any]] = []
    for row in reader:
        cleaned = {k: _coerce_value(v) for k, v in row.items() if k is not None}
        rows.append(cleaned)

    return DocumentContent(
        id=f"doc-{uuid.uuid4().hex[:12]}",
        filename=filename,
        content_type="text/csv",
        row_count=len(rows),
        column_names=list(columns),
        rows=rows,
    )


def parse_excel(
    filename: str,
    content: bytes,
    *,
    extra_skip_patterns: tuple[str, ...] = (),
    section_labels: frozenset[str] = frozenset(),
) -> DocumentContent:
    """Parse an Excel file (.xlsx / .xls) into a Document.

    Handles report-style exports that prepend several rows of metadata
    before the real column header.  The parser scans forward to find the
    first row that looks like a genuine header.

    *section_labels* are domain-specific labels (e.g. ``"current"``,
    ``"vacant"``) that identify section context rows in hierarchical
    reports.  When empty, section detection is skipped.

    Requires the openpyxl optional dependency.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required for Excel parsing. Install with: pip install remi[documents]"
        ) from exc

    content_type_xlsx = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return DocumentContent(
            id=f"doc-{uuid.uuid4().hex[:12]}",
            filename=filename,
            content_type=content_type_xlsx,
        )

    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        return DocumentContent(
            id=f"doc-{uuid.uuid4().hex[:12]}",
            filename=filename,
            content_type=content_type_xlsx,
        )

    header_idx = _find_header_row(all_rows, extra_skip_patterns=extra_skip_patterns)
    meta = _extract_metadata(all_rows, header_idx)
    header = all_rows[header_idx]
    columns = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(header)]

    rows, _section_ctx = _parse_data_rows(
        all_rows[header_idx + 1 :], columns, section_labels=section_labels
    )

    return DocumentContent(
        id=f"doc-{uuid.uuid4().hex[:12]}",
        filename=filename,
        content_type=content_type_xlsx,
        row_count=len(rows),
        column_names=columns,
        rows=rows,
        metadata=meta,
    )


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

_DO_NOT_USE_RE = re.compile(r"^do\s+not\s+use\s*[-–]?\s*", re.I)

# "136 Units", "17 Units" — section summary rows in rent roll reports.
_UNITS_SUMMARY_RE = re.compile(r"^\d+\s+units?$", re.I)

# "Mar 2026", "Jun 2026" — month/year section labels in lease expiration reports.
_MONTH_YEAR_RE = re.compile(
    r"^(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{4}$", re.I,
)


def _normalize_cell_address(val: str) -> str:
    """Strip 'DO NOT USE' prefixes from address strings."""
    return _DO_NOT_USE_RE.sub("", val).strip()


def _is_section_label(val: str, section_labels: frozenset[str]) -> bool:
    """True when a single-cell value is a section label, not a property address."""
    lower = val.lower()
    if section_labels and lower in section_labels:
        return True
    if _MONTH_YEAR_RE.match(val):
        return True
    return False


def _parse_data_rows(
    raw_rows: list[tuple[Any, ...]],
    columns: list[str],
    *,
    section_labels: frozenset[str] = frozenset(),
) -> tuple[list[dict[str, Any]], dict[str, str]]:
    """Convert raw sheet rows into data row dicts with section headers denormalized.

    Hierarchical reports (rent roll, delinquency, lease expiration) embed
    context rows between data rows — a row with a single non-null value
    that identifies either:
      - A section label (matched against *section_labels*) → ``_ctx_section_label``
      - A free-text value (e.g. address)                   → ``_ctx_property_address``

    When *section_labels* is empty, single-cell rows are still captured as
    ``_ctx_property_address`` context (address heuristic).

    **Denormalization:** Every data row gets a ``_section_property`` column
    containing the last-seen property address from section headers. This
    means downstream consumers always see a flat table where every row has
    its property — no need to track section headers separately.

    Returns (data_rows, final_context).
    """
    context: dict[str, str] = {}
    rows: list[dict[str, Any]] = []
    n_cols = len(columns)

    for raw_row in raw_rows:
        # Skip fully empty rows
        non_null = [(i, c) for i, c in enumerate(raw_row) if c is not None]
        if not non_null:
            continue

        # Coerce all cells
        coerced = [_coerce_value(c) for c in raw_row]
        non_empty = [(i, v) for i, v in enumerate(coerced) if v is not None]

        # Section context row: exactly one non-null cell across all columns
        if len(non_empty) == 1 and n_cols > 1:
            raw_val = str(non_empty[0][1]).strip()
            clean_val = _normalize_cell_address(raw_val)

            # "136 Units" etc. — section boundary summaries, not context.
            if _UNITS_SUMMARY_RE.match(clean_val):
                continue

            if _is_section_label(clean_val, section_labels):
                context["_ctx_section_label"] = clean_val
            else:
                context["_ctx_property_address"] = clean_val
                context.pop("_ctx_section_label", None)
            continue

        # Data row — build dict and attach current context
        row_dict: dict[str, Any] = {}
        for col_name, val in zip(columns, coerced, strict=False):
            if isinstance(val, str):
                val = _normalize_cell_address(val) if _DO_NOT_USE_RE.match(val) else val
            row_dict[col_name] = val
        row_dict.update(context)

        # Denormalize: inject the current section property address so every
        # row carries its property — consumers never need to track headers.
        if "_ctx_property_address" in context:
            row_dict["_section_property"] = context["_ctx_property_address"]

        rows.append(row_dict)

    return rows, context


def _compile_skip_pattern(extra: tuple[str, ...] = ()) -> re.Pattern[str]:
    """Compile metadata-skip regex from generic + domain-specific patterns."""
    all_parts = list(_GENERIC_METADATA_PATTERNS) + list(extra)
    return re.compile(r"^(" + "|".join(all_parts) + r")", re.I)


_METADATA_KV_RE = re.compile(r"^([A-Za-z][A-Za-z0-9 ]+?):\s*(.+)$")


def _extract_metadata(rows: list[tuple[Any, ...]], header_idx: int) -> dict[str, str]:
    """Extract key-value metadata from the rows above the column header.

    Report-style exports often prepend lines like:
      ``Report Group: Acme Corp``        ← single cell, "key: value"
      ``Report Group | Acme Corp``       ← two cells, key in col A, value in col B
      ``Exported On: 03/23/2026 11:54 AM``

    Returns a dict of normalised keys → raw values.
    """
    meta: dict[str, str] = {}
    for row in rows[:header_idx]:
        non_null = [c for c in row if c is not None]
        if not non_null:
            continue

        if len(non_null) == 1:
            # Single-cell row — may be "Key: Value" or a bare title
            text = str(non_null[0]).strip()
            m = _METADATA_KV_RE.match(text)
            if m:
                key = m.group(1).strip().lower().replace(" ", "_")
                meta[key] = m.group(2).strip()
            elif not meta and text:
                meta["report_title"] = text

        elif len(non_null) == 2:
            # Two-cell row — label in first non-null cell, value in second.
            # AppFolio sometimes exports metadata this way.
            label = str(non_null[0]).strip()
            value = str(non_null[1]).strip()
            if label and value and re.match(r"^[A-Za-z][A-Za-z0-9 ]+$", label):
                key = label.lower().replace(" ", "_")
                meta[key] = value

    return meta


def _find_header_row(
    rows: list[tuple[Any, ...]],
    *,
    extra_skip_patterns: tuple[str, ...] = (),
) -> int:
    """Return the index of the first row that looks like a real column header."""
    pattern = _compile_skip_pattern(extra_skip_patterns)
    for idx, row in enumerate(rows):
        non_null = [c for c in row if c is not None]
        if len(non_null) < _MIN_HEADER_CELLS:
            continue

        first_val = str(non_null[0]).strip() if non_null else ""

        if pattern.match(first_val):
            continue

        if isinstance(non_null[0], (_dt.datetime, _dt.date)):
            continue

        numeric_count = sum(1 for c in non_null if isinstance(c, (int, float)))
        if numeric_count > len(non_null) // 2:
            continue

        return idx

    return 0


def _coerce_value(val: Any) -> Any:
    """Best-effort coercion: numbers, booleans, nulls."""
    if val is None:
        return None
    if not isinstance(val, str):
        return val
    stripped = val.strip()
    if stripped == "":
        return None
    if stripped.lower() in ("true", "yes"):
        return True
    if stripped.lower() in ("false", "no"):
        return False
    try:
        return int(stripped)
    except ValueError:
        pass
    try:
        return float(stripped)
    except ValueError:
        pass
    return stripped
